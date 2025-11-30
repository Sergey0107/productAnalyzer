import time
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook
import xlrd

from handlers.file_handler import FileHandler


class XlsHandler(FileHandler):

    def get_data_from_xls_file(self, file_path: Path) -> List[Dict[str, Any]]:
        start_time = time.time()

        file_extension = file_path.suffix.lower()

        if file_extension == '.xls':
            return self._parse_xls(file_path, start_time)
        else:
            return self._parse_xlsx(file_path, start_time)

    def _parse_xls(self, file_path: Path, start_time: float) -> List[Dict[str, Any]]:
        workbook = xlrd.open_workbook(file_path)
        sheets_data = []
        total_rows = 0

        for sheet in workbook.sheets():
            sheet_rows = []

            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                row_cells = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in row
                ]
                # Пропускаем полностью пустые строки
                if any(cell for cell in row_cells):
                    sheet_rows.append(row_cells)
                    total_rows += 1

            sheets_data.append({
                "sheet_name": sheet.name,
                "rows": sheet_rows
            })

        elapsed_time = time.time() - start_time
        print(f"[DEBUG] XLS парсинг (xlrd) | sheets={len(workbook.sheets())} | rows={total_rows} | time={elapsed_time:.2f}s")

        return sheets_data

    def _parse_xlsx(self, file_path: Path, start_time: float) -> List[Dict[str, Any]]:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheets_data = []
        total_rows = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):

                row_cells = [
                    str(cell).strip() if cell is not None else ""
                    for cell in row
                ]
                # Пропускаем полностью пустые строки
                if any(cell for cell in row_cells):
                    sheet_rows.append(row_cells)
                    total_rows += 1

            sheets_data.append({
                "sheet_name": sheet_name,
                "rows": sheet_rows
            })

        workbook.close()

        elapsed_time = time.time() - start_time
        print(f"[DEBUG] XLSX парсинг (openpyxl) | sheets={len(workbook.sheetnames)} | rows={total_rows} | time={elapsed_time:.2f}s")

        return sheets_data